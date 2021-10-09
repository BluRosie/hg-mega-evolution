
from glob import glob
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
import itertools
import hashlib
import subprocess
import sys

def getpmindex(chine:str,table:[str]):
    txt = open("data/text/238.txt","r",encoding="utf-8")
    s = 0
    txt2 = txt.readlines()
    for i in txt2 :
        if chine == i.replace("\n",'').replace(" ",''):
            break
        s += 1
    txt.close()
    return table[s].split(" equ")[0]

txt = open("armips/include/monnums.s","r",encoding="utf-8")
table = txt.readlines()
txt.close()

the = load_workbook('data/wildmon.xlsx')#加载xlsx
the.guess_types = True

sheets = the.sheetnames
wa = the[sheets[0]]
wb = the[sheets[1]]
wc = the[sheets[2]]

asm_file = "armips/data/wildmon.s"
i = 0

while True:
    if wb["A" + str(i+3)].value == None:
        break
    i+=1

numof_wild = i
i = 0
word = ".nds\n\
.thumb\n\n\
.include \"armips/include/macros.s\"\n\
.include \"armips/include/monnums.s\"\n\n"

while i < numof_wild:
    word += "wilddata " + str(i) + "\n"
    word += "    wildrate " + str(wa["B" + str(i * 6 + 3)].value) + ", " + str(wb["F" + str(i + 3)].value) + ", " + str(wc["T" + str((i + 1) * 3)].value) + ", "\
         + str(wc["B" + str((i + 1) * 3)].value) + ", " + str(wc["C" + str((i + 1) * 3)].value) + ", " + str(wc["D" + str((i + 1) * 3)].value)

    word += "\n    walk_lv " + str(wa["C"+ str(i * 6 + 3)].value) + ", " + str(wa["G"+ str(i * 6 + 3)].value) + ", " + str(wa["K"+ str(i * 6 + 3)].value) + ", " \
        + str(wa["O"+ str(i * 6 + 3)].value) + ", " + str(wa["S"+ str(i * 6 + 3)].value) + ", " + str(wa["W"+ str(i * 6 + 3)].value) + ", "\
        + str(wa["C"+ str(i * 6 + 6)].value) + ", " + str(wa["G"+ str(i * 6 + 6)].value) + ", " + str(wa["K"+ str(i * 6 + 6)].value) + ", " \
        + str(wa["O"+ str(i * 6 + 6)].value) + ", " + str(wa["S"+ str(i * 6 + 6)].value) + ", " + str(wa["W"+ str(i * 6 + 6)].value) 
    
    word += "\n    morning_poke " + getpmindex(wa["D"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["H"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["L"+ str(i * 6 + 3)].value,table) + ", "\
        + getpmindex(wa["P"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["T"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["X"+ str(i * 6 + 3)].value,table) + ", "\
        + getpmindex(wa["D"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["H"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["L"+ str(i * 6 + 6)].value,table) + ", "\
        + getpmindex(wa["P"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["T"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["X"+ str(i * 6 + 6)].value,table)
    
    word += "\n    day_poke " + getpmindex(wa["E"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["I"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["M"+ str(i * 6 + 3)].value,table) + ", "\
        + getpmindex(wa["Q"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["U"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["Y"+ str(i * 6 + 3)].value,table) + ", "\
        + getpmindex(wa["E"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["I"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["M"+ str(i * 6 + 6)].value,table) + ", "\
        + getpmindex(wa["Q"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["U"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["Y"+ str(i * 6 + 6)].value,table)

    word += "\n    night_poke " + getpmindex(wa["F"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["J"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["N"+ str(i * 6 + 3)].value,table) + ", "\
        + getpmindex(wa["R"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["V"+ str(i * 6 + 3)].value,table) + ", " + getpmindex(wa["Z"+ str(i * 6 + 3)].value,table) + ", "\
        + getpmindex(wa["F"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["J"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["N"+ str(i * 6 + 6)].value,table) + ", "\
        + getpmindex(wa["R"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["V"+ str(i * 6 + 6)].value,table) + ", " + getpmindex(wa["Z"+ str(i * 6 + 6)].value,table)
    
    word += "\n    pokegear " + getpmindex(wb["B"+ str(i + 3)].value,table) + ", " + getpmindex(wb["C"+ str(i + 3)].value,table) + ", " + getpmindex(wb["D"+ str(i + 3)].value,table) + ", " + getpmindex(wb["E"+ str(i + 3)].value,table)

    word += "\n    wildmon " + str(wb["G"+ str(i+ 3)].value) + ", " + str(wb["H"+ str(i+ 3)].value) + ", " + getpmindex(wb["I"+ str(i + 3)].value,table)
    word += "\n    wildmon " + str(wb["J"+ str(i+ 3)].value) + ", " + str(wb["K"+ str(i+ 3)].value) + ", " + getpmindex(wb["L"+ str(i + 3)].value,table)
    word += "\n    wildmon " + str(wb["M"+ str(i+ 3)].value) + ", " + str(wb["N"+ str(i+ 3)].value) + ", " + getpmindex(wb["O"+ str(i + 3)].value,table)
    word += "\n    wildmon " + str(wb["P"+ str(i+ 3)].value) + ", " + str(wb["Q"+ str(i+ 3)].value) + ", " + getpmindex(wb["R"+ str(i + 3)].value,table)
    word += "\n    wildmon " + str(wb["S"+ str(i+ 3)].value) + ", " + str(wb["T"+ str(i+ 3)].value) + ", " + getpmindex(wb["U"+ str(i + 3)].value,table)

    word += "\n    wildmon " + str(wc["U"+ str((i + 1) * 3)].value) + ", " + str(wc["V"+ str((i + 1) * 3)].value) + ", " + getpmindex(wc["W"+ str((i + 1) * 3)].value,table)
    word += "\n    wildmon " + str(wc["X"+ str((i + 1) * 3)].value) + ", " + str(wc["Y"+ str((i + 1) * 3)].value) + ", " + getpmindex(wc["Z"+ str((i + 1) * 3)].value,table)

    word += "\n    wildmon " + str(wc["E"+ str((i + 1) * 3)].value) + ", " + str(wc["F"+ str((i + 1) * 3)].value) + ", " + getpmindex(wc["G"+ str((i + 1) * 3)].value,table)
    word += "\n    wildmon " + str(wc["H"+ str((i + 1) * 3)].value) + ", " + str(wc["I"+ str((i + 1) * 3)].value) + ", " + getpmindex(wc["J"+ str((i + 1) * 3)].value,table)
    word += "\n    wildmon " + str(wc["K"+ str((i + 1) * 3)].value) + ", " + str(wc["L"+ str((i + 1) * 3)].value) + ", " + getpmindex(wc["M"+ str((i + 1) * 3)].value,table)
    word += "\n    wildmon " + str(wc["N"+ str((i + 1) * 3)].value) + ", " + str(wc["O"+ str((i + 1) * 3)].value) + ", " + getpmindex(wc["P"+ str((i + 1) * 3)].value,table)
    word += "\n    wildmon " + str(wc["Q"+ str((i + 1) * 3)].value) + ", " + str(wc["R"+ str((i + 1) * 3)].value) + ", " + getpmindex(wc["S"+ str((i + 1) * 3)].value,table)

    word += "\n    wildmon " + str(wc["E"+ str((i + 1) * 3 + 1)].value) + ", " + str(wc["F"+ str((i + 1) * 3 + 1)].value) + ", " + getpmindex(wc["G"+ str((i + 1) * 3 + 1)].value,table)
    word += "\n    wildmon " + str(wc["H"+ str((i + 1) * 3 + 1)].value) + ", " + str(wc["I"+ str((i + 1) * 3 + 1)].value) + ", " + getpmindex(wc["J"+ str((i + 1) * 3 + 1)].value,table)
    word += "\n    wildmon " + str(wc["K"+ str((i + 1) * 3 + 1)].value) + ", " + str(wc["L"+ str((i + 1) * 3 + 1)].value) + ", " + getpmindex(wc["M"+ str((i + 1) * 3 + 1)].value,table)
    word += "\n    wildmon " + str(wc["N"+ str((i + 1) * 3 + 1)].value) + ", " + str(wc["O"+ str((i + 1) * 3 + 1)].value) + ", " + getpmindex(wc["P"+ str((i + 1) * 3 + 1)].value,table)
    word += "\n    wildmon " + str(wc["Q"+ str((i + 1) * 3 + 1)].value) + ", " + str(wc["R"+ str((i + 1) * 3 + 1)].value) + ", " + getpmindex(wc["S"+ str((i + 1) * 3 + 1)].value,table)

    word += "\n    wildmon " + str(wc["E"+ str((i + 1) * 3 + 2)].value) + ", " + str(wc["F"+ str((i + 1) * 3 + 2)].value) + ", " + getpmindex(wc["G"+ str((i + 1) * 3 + 2)].value,table)
    word += "\n    wildmon " + str(wc["H"+ str((i + 1) * 3 + 2)].value) + ", " + str(wc["I"+ str((i + 1) * 3 + 2)].value) + ", " + getpmindex(wc["J"+ str((i + 1) * 3 + 2)].value,table)
    word += "\n    wildmon " + str(wc["K"+ str((i + 1) * 3 + 2)].value) + ", " + str(wc["L"+ str((i + 1) * 3 + 2)].value) + ", " + getpmindex(wc["M"+ str((i + 1) * 3 + 2)].value,table)
    word += "\n    wildmon " + str(wc["N"+ str((i + 1) * 3 + 2)].value) + ", " + str(wc["O"+ str((i + 1) * 3 + 2)].value) + ", " + getpmindex(wc["P"+ str((i + 1) * 3 + 2)].value,table)
    word += "\n    wildmon " + str(wc["Q"+ str((i + 1) * 3 + 2)].value) + ", " + str(wc["R"+ str((i + 1) * 3 + 2)].value) + ", " + getpmindex(wc["S"+ str((i + 1) * 3 + 2)].value,table)

    word += "\n    swarm " + getpmindex(wb["V"+ str(i + 3)].value,table) + ", " + getpmindex(wb["W"+ str(i + 3)].value,table) + ", " + getpmindex(wb["X"+ str(i + 3)].value,table) + ", " + getpmindex(wb["Y"+ str(i + 3)].value,table)

    word += "\n\n"
    i +=1

file = open(asm_file,"w",encoding="utf-8")
file.write(word)
file.close()