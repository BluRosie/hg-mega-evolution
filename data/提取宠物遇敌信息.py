
from glob import glob
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl import load_workbook
import shutil
import itertools
import hashlib
import subprocess
import sys

DIR = "data/a037/"
get_dir = os.listdir(DIR)
wb = load_workbook('data/wildmon.xlsx')#加载xlsx
wb.guess_types = True

sheets = wb.sheetnames
ws = wb[sheets[0]]
wa = wb[sheets[1]]
wc = wb[sheets[2]]
wf = wb[sheets[3]]

#print(wb.sheetnames)#列表名字
s = 0

for i in get_dir:
    with open(DIR+i,'rb') as rom:
        sell = str(s * 6 +3)
        ws["A" + sell] = str(s)
        wa["A" + str(s+3)] = str(s)
        wc["A" + str(s * 3+3)] = str(s)
        ws["B" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        wa["F" + str(s + 3)] = int(rom.read(1)[0])
        
        wc["T" + str(s * 3+3)] = int(rom.read(1)[0])
        wc["B" + str(s * 3+3)] = int(rom.read(1)[0])
        wc["C" + str(s * 3+3)] = int(rom.read(1)[0])
        wc["D" + str(s * 3+3)] = int(rom.read(1)[0])
        rom.seek(0x8)
        ws["C" + sell] = int(rom.read(1)[0])
        ws["G" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        ws["K" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        ws["O" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        ws["S" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        ws["W" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        sell = str(s * 6 +6)
        ws["C" + sell] = int(rom.read(1)[0])
        ws["G" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        ws["K" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        ws["O" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        ws["S" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)
        ws["W" + sell] = int().from_bytes(rom.read(1), byteorder='big', signed=True)

        ########################清晨#############
        sell = str(s * 6 +3) 
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["D" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["H" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["L" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["P" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["T" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["X" + sell] = wf["A" + MON].value

        sell = str(s * 6 +6)

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["D" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["H" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["L" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["P" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["T" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["X" + sell] = wf["A" + MON].value

        ########################中午#############
        sell = str(s * 6 +3)
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["E" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["I" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["M" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["Q" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["U" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["Y" + sell] = wf["A" + MON].value

        sell = str(s * 6 +6)

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["E" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["I" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["M" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["Q" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["U" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["Y" + sell] = wf["A" + MON].value

         ########################晚上#############
        sell = str(s * 6 +3)
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["F" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["J" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["N" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["R" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["V" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["Z" + sell] = wf["A" + MON].value

        sell = str(s * 6 +6)

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["F" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["J" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["N" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["R" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["V" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        ws["Z" + sell] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["B" + str(s + 3)] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["C" + str(s + 3)] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["D" + str(s + 3)] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["E" + str(s + 3)] = wf["A" + MON].value

    #########################################################
        wa["G" + str(s + 3)] = int(rom.read(1)[0])
        wa["H" + str(s + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["I" + str(s + 3)] = wf["A" + MON].value

        wa["J" + str(s + 3)] = int(rom.read(1)[0])
        wa["K" + str(s + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["L" + str(s + 3)] = wf["A" + MON].value

        wa["M" + str(s + 3)] = int(rom.read(1)[0])
        wa["N" + str(s + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["O" + str(s + 3)] = wf["A" + MON].value

        wa["P" + str(s + 3)] = int(rom.read(1)[0])
        wa["Q" + str(s + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["R" + str(s + 3)] = wf["A" + MON].value

        wa["S" + str(s + 3)] = int(rom.read(1)[0])
        wa["T" + str(s + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["U" + str(s + 3)] = wf["A" + MON].value
#################################################################
        wc["U" + str(s*3 + 3)] = int(rom.read(1)[0])
        wc["V" + str(s*3  + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["W" + str(s*3  + 3)] = wf["A" + MON].value

        wc["X" + str(s*3  + 3)] = int(rom.read(1)[0])
        wc["Y" + str(s*3  + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["Z" + str(s*3  + 3)] = wf["A" + MON].value
    ##########################################
        wc["E" + str(s*3 + 3)] = int(rom.read(1)[0])
        wc["F" + str(s*3  + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["G" + str(s*3  + 3)] = wf["A" + MON].value

        wc["H" + str(s*3 + 3)] = int(rom.read(1)[0])
        wc["I" + str(s*3  + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["J" + str(s*3  + 3)] = wf["A" + MON].value

        wc["K" + str(s*3 + 3)] = int(rom.read(1)[0])
        wc["L" + str(s*3  + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["M" + str(s*3  + 3)] = wf["A" + MON].value

        wc["N" + str(s*3 + 3)] = int(rom.read(1)[0])
        wc["O" + str(s*3  + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["P" + str(s*3  + 3)] = wf["A" + MON].value

        wc["Q" + str(s*3 + 3)] = int(rom.read(1)[0])
        wc["R" + str(s*3  + 3)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["S" + str(s*3  + 3)] = wf["A" + MON].value

        #########
        wc["E" + str(s*3 + 4)] = int(rom.read(1)[0])
        wc["F" + str(s*3  + 4)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["G" + str(s*3  + 4)] = wf["A" + MON].value

        wc["H" + str(s*3 + 4)] = int(rom.read(1)[0])
        wc["I" + str(s*3  + 4)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["J" + str(s*3  + 4)] = wf["A" + MON].value

        wc["K" + str(s*3 + 4)] = int(rom.read(1)[0])
        wc["L" + str(s*3  + 4)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["M" + str(s*3  + 4)] = wf["A" + MON].value

        wc["N" + str(s*3 + 4)] = int(rom.read(1)[0])
        wc["O" + str(s*3  + 4)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["P" + str(s*3  + 4)] = wf["A" + MON].value

        wc["Q" + str(s*3 + 4)] = int(rom.read(1)[0])
        wc["R" + str(s*3  + 4)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["S" + str(s*3  + 4)] = wf["A" + MON].value
    ##########
        wc["E" + str(s*3 + 5)] = int(rom.read(1)[0])
        wc["F" + str(s*3  + 5)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["G" + str(s*3  + 5)] = wf["A" + MON].value

        wc["H" + str(s*3 + 5)] = int(rom.read(1)[0])
        wc["I" + str(s*3  + 5)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["J" + str(s*3  + 5)] = wf["A" + MON].value

        wc["K" + str(s*3 + 5)] = int(rom.read(1)[0])
        wc["L" + str(s*3  + 5)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["M" + str(s*3  + 5)] = wf["A" + MON].value

        wc["N" + str(s*3 + 5)] = int(rom.read(1)[0])
        wc["O" + str(s*3  + 5)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["P" + str(s*3  + 5)] = wf["A" + MON].value

        wc["Q" + str(s*3 + 5)] = int(rom.read(1)[0])
        wc["R" + str(s*3  + 5)] = int(rom.read(1)[0])
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wc["S" + str(s*3  + 5)] = wf["A" + MON].value
    
    ####################
        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["V" + str(s+3)] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["W" + str(s+3 )] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["X" + str(s+3 )] = wf["A" + MON].value

        mon = rom.read(2)
        MON = str((mon[1] << 8 | mon[0])  + 1)
        wa["Y" + str(s+3 )] = wf["A" + MON].value
        rom.close()
    s+=1
wb.save("data/w.xlsx")